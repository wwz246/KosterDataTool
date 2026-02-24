from __future__ import annotations

import copy
import re
from pathlib import Path

from openpyxl import Workbook

from .colmap import _append_run_report
from .curve_export import export_cv_block, export_eis_block, export_gcd_block
from .excel_writer import FLOAT2_FMT, INT_FMT, blank_col_after, write_block
from .export_blocks import Block3Header
from .gcd_window_metrics import compute_gcd_file_metrics
from .rate_retention import build_rate_and_retention_for_battery


NUM_RE = re.compile(r"^(CV|GCD|EIS)-([+-]?\d+(?:\.\d+)?)\.txt$", re.IGNORECASE)


def _num_text(path: str, prefix: str) -> str:
    m = NUM_RE.match(Path(path).name)
    if not m or m.group(1).upper() != prefix.upper():
        return ""
    return m.group(2)


def _find_file(files, num: float):
    for f in files:
        if abs(float(f.num) - float(num)) < 1e-12:
            return f.path
    return None


def _empty_curve_block() -> Block3Header:
    return Block3Header(h1=["X", "Y"], h2=["", ""], h3=["", ""], data=[[], []], warnings=[])


def _electrode_curve_block(block: Block3Header, battery_name: str) -> Block3Header:
    out = copy.deepcopy(block)
    if len(out.h3) >= 2:
        out.h3[1] = battery_name
    return out




def _record_failure(run_report_path: str, logger, file_path: str, exc: Exception, code: str = "E9001") -> None:
    msg = f"{code} 文件失败 file={file_path} err={exc}"
    logger.error(msg, code=code, file_path=file_path, error=str(exc))
    _append_run_report(run_report_path, msg)


def _apply_param_cycle_formats(ws, start_row: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in (4, 5, 8, 9, 10):
            ws.cell(row=r, column=c).number_format = FLOAT2_FMT
        for c in (6, 7):
            ws.cell(row=r, column=c).number_format = INT_FMT

def _build_rate_retention_blocks(battery, params, logger, run_report_path: str, compact_rate_columns: bool):
    bparam = params["battery_params"].get(battery.name, {})
    root_params = {
        "a_geom": params.get("a_geom", 1.0),
        "v_start": bparam.get("v_start"),
        "v_end": bparam.get("v_end"),
        "k_factor": bparam.get("k"),
    }
    return build_rate_and_retention_for_battery(
        gcd_files=[f.path for f in battery.gcd_files],
        n_gcd=int(bparam.get("n_gcd", 1)),
        output_type=params.get("output_type", "Csp"),
        root_params=root_params,
        battery_params=bparam,
        logger=logger,
        run_report_path=run_report_path,
        csp_column_choice=params.get("electrode_rate_csp_column"),
        compact_rate_columns=compact_rate_columns,
    )


def build_electrode_workbook(scan_result, selections, params, logger, run_report_path) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rate"

    selected_bats = sorted([b for b in scan_result.batteries if b.name in set(selections.get("batteries", []))], key=lambda x: x.name)

    col = 1
    end_row = 1
    for b in selected_bats:
        try:
            rr = _build_rate_retention_blocks(b, params, logger, run_report_path, compact_rate_columns=True)
            if len(rr.rate.h3) >= 2:
                rr.rate.h3[0] = ""
                for i in range(1, len(rr.rate.h3)):
                    rr.rate.h3[i] = b.name
        except Exception as exc:
            _record_failure(run_report_path, logger, b.name, exc)
            rr = type("Tmp", (), {"rate": _empty_curve_block(), "retention": _empty_curve_block()})()
        ec, er = write_block(ws, col, 1, rr.rate)
        col = ec
        end_row = max(end_row, er)

    for n in selections.get("cv_nums", []):
        ws_cv = wb.create_sheet(f"CV-{n}")
        col = 1
        for b in selected_bats:
            fp = _find_file(b.cv_files, float(n))
            if not fp:
                continue
            bp = params["battery_params"][b.name]
            try:
                blk = export_cv_block(fp, int(bp.get("n_cv", 1)), float(params.get("a_geom", 1.0)), float(bp.get("m_pos", 0.0)), float(bp.get("m_neg", 0.0)), float(bp.get("p_active", 100.0)), logger, run_report_path)
            except Exception as exc:
                _record_failure(run_report_path, logger, fp, exc)
                continue
            blk = _electrode_curve_block(blk, b.name)
            write_block(ws_cv, col, 1, blk)
            col += len(blk.h1)

    for n in selections.get("gcd_nums", []):
        ws_g = wb.create_sheet(f"GCD-{n}")
        col = 1
        for b in selected_bats:
            fp = _find_file(b.gcd_files, float(n))
            if not fp:
                continue
            bp = params["battery_params"][b.name]
            try:
                blk = export_gcd_block(fp, int(bp.get("n_gcd", 1)), logger, run_report_path)
            except Exception as exc:
                _record_failure(run_report_path, logger, fp, exc)
                continue
            blk = _electrode_curve_block(blk, b.name)
            write_block(ws_g, col, 1, blk)
            col += len(blk.h1)

    for n in selections.get("eis_nums", []):
        ws_e = wb.create_sheet(f"EIS-{n}")
        col = 1
        for b in selected_bats:
            fp = _find_file(b.eis_files, float(n))
            if not fp:
                continue
            try:
                blk = export_eis_block(fp, float(params.get("a_geom", 1.0)), logger, run_report_path)
            except Exception as exc:
                _record_failure(run_report_path, logger, fp, exc)
                continue
            blk = _electrode_curve_block(blk, b.name)
            write_block(ws_e, col, 1, blk)
            col += len(blk.h1)

    return wb


def _build_param_summary_sheet(wb: Workbook, scan_result, params, logger, run_report_path: str):
    ws = wb.create_sheet("参数汇总")
    cols = ["电池名", "m_pos(mg)", "m_neg(mg)", "p_active(%)", "K(—)", "N_CV", "N_GCD", "V_start(V)", "V_end(V)", "CV最大圈数", "GCD最大圈数", "输出类型(Csp/Qsp)", "主顺序"]
    for c, name in enumerate(cols, start=1):
        ws.cell(row=1, column=c, value=name)
    row = 2
    all_cycle_rows: list[list] = []
    for b in sorted(scan_result.batteries, key=lambda x: x.name):
        bp = params["battery_params"][b.name]
        ws.cell(row=row, column=1, value=b.name)
        ws.cell(row=row, column=2, value=bp.get("m_pos"))
        ws.cell(row=row, column=3, value=bp.get("m_neg"))
        ws.cell(row=row, column=4, value=bp.get("p_active"))
        ws.cell(row=row, column=5, value=bp.get("k") if params.get("output_type") == "Csp" else "")
        ws.cell(row=row, column=6, value=bp.get("n_cv"))
        ws.cell(row=row, column=7, value=bp.get("n_gcd"))
        ws.cell(row=row, column=8, value=bp.get("v_start"))
        ws.cell(row=row, column=9, value=bp.get("v_end"))
        ws.cell(row=row, column=10, value=b.cv_max_cycle)
        ws.cell(row=row, column=11, value=b.gcd_max_cycle)
        ws.cell(row=row, column=12, value=params.get("output_type"))
        ws.cell(row=row, column=13, value=bp.get("main_order", ""))
        row += 1

        for g in b.gcd_files:
            try:
                gm = compute_gcd_file_metrics(
                    file_path=g.path,
                    root_params={"v_start": bp.get("v_start"), "v_end": bp.get("v_end"), "a_geom": params.get("a_geom", 1.0), "output_type": params.get("output_type"), "k_factor": bp.get("k"), "n_gcd": bp.get("n_gcd", 1)},
                    battery_params=bp,
                    logger=logger,
                    run_report_path=run_report_path,
                )
            except Exception as exc:
                _record_failure(run_report_path, logger, g.path, exc)
                gm = type("Tmp", (), {"cycles": {}})()
            for k in range(1, int(b.gcd_max_cycle or 0) + 1):
                rep = gm.cycles.get(k)
                m_active_g = max(1e-12, (float(bp.get("m_pos", 0.0)) + float(bp.get("m_neg", 0.0))) * float(bp.get("p_active", 100.0)) / 100000.0)
                if rep is None:
                    all_cycle_rows.append([b.name, _num_text(g.path, "GCD"), k, "NA", "NA", "NA", "NA", "NA", "NA", "NA"])
                    continue
                ce = (100.0 * rep.delta_q_dis / rep.delta_q_chg) if (rep.delta_q_chg is not None and rep.delta_q_dis is not None and rep.delta_q_chg != 0) else None
                qsp_chg = (rep.delta_q_chg / m_active_g) if rep.delta_q_chg is not None else None
                qsp_dis = (rep.delta_q_dis / m_active_g) if rep.delta_q_dis is not None else None
                kf = float(bp.get("k", 1.0))
                csp_chg_eff = ((rep.delta_q_eff_chg * 3.6) / (rep.delta_v_eff_chg * m_active_g) * kf) if (rep.delta_q_eff_chg is not None and rep.delta_v_eff_chg and rep.delta_v_eff_chg > 0) else None
                csp_dis_eff = ((rep.delta_q_eff_dis * 3.6) / (rep.delta_v_eff_dis * m_active_g) * kf) if (rep.delta_q_eff_dis is not None and rep.delta_v_eff_dis and rep.delta_v_eff_dis > 0) else None
                all_cycle_rows.append([
                    b.name,
                    _num_text(g.path, "GCD"),
                    k,
                    qsp_chg,
                    qsp_dis,
                    csp_chg_eff,
                    csp_dis_eff,
                    ce,
                    rep.r_drop,
                    rep.r_turn,
                ])

    header2 = ["电池名", "J_label(A/g)", "圈号k", "Qsp_chg", "Qsp_dis", "Csp_chg_eff", "Csp_dis_eff", "CE", "R↓", "R_turn"]
    start = row + 5
    for c, h in enumerate(header2, start=1):
        ws.cell(row=start, column=c, value=h)
    for i, vals in enumerate(all_cycle_rows, start=1):
        for c, v in enumerate(vals, start=1):
            ws.cell(row=start + i, column=c, value=v)
    _apply_param_cycle_formats(ws, start + 1, start + len(all_cycle_rows))


def build_battery_workbook(scan_result, params, logger, run_report_path) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _build_param_summary_sheet(wb, scan_result, params, logger, run_report_path)

    for b in sorted(scan_result.batteries, key=lambda x: x.name):
        ws = wb.create_sheet(b.name)
        bp = params["battery_params"][b.name]
        col = 1

        for f in sorted(b.cv_files, key=lambda x: x.num):
            try:
                blk = export_cv_block(f.path, int(bp.get("n_cv", 1)), float(params.get("a_geom", 1.0)), float(bp.get("m_pos", 0.0)), float(bp.get("m_neg", 0.0)), float(bp.get("p_active", 100.0)), logger, run_report_path)
            except Exception as exc:
                _record_failure(run_report_path, logger, f.path, exc)
                continue
            write_block(ws, col, 1, blk)
            col += len(blk.h1)
        col = blank_col_after(ws, col)

        for f in sorted(b.gcd_files, key=lambda x: x.num):
            try:
                blk = export_gcd_block(f.path, int(bp.get("n_gcd", 1)), logger, run_report_path)
            except Exception as exc:
                _record_failure(run_report_path, logger, f.path, exc)
                continue
            write_block(ws, col, 1, blk)
            col += len(blk.h1)
        col = blank_col_after(ws, col)

        for f in sorted(b.eis_files, key=lambda x: x.num):
            try:
                blk = export_eis_block(f.path, float(params.get("a_geom", 1.0)), logger, run_report_path)
            except Exception as exc:
                _record_failure(run_report_path, logger, f.path, exc)
                continue
            write_block(ws, col, 1, blk)
            col += len(blk.h1)
        col = blank_col_after(ws, col)

        try:
            rr = _build_rate_retention_blocks(b, params, logger, run_report_path, compact_rate_columns=False)
        except Exception as exc:
            _record_failure(run_report_path, logger, b.name, exc)
            rr = type("Tmp", (), {"rate": _empty_curve_block(), "retention": _empty_curve_block()})()
        rate = copy.deepcopy(rr.rate)
        retention = copy.deepcopy(rr.retention)
        if len(rate.h3) >= 2:
            rate.h3 = [""] + ["Rate"] + ["" for _ in rate.h3[2:]]
        if len(retention.h3) >= 2:
            retention.h3 = [""] + ["Retention"] + ["" for _ in retention.h3[2:]]
        _, er = write_block(ws, col, 1, rate)
        write_block(ws, col, er + 2, retention)

    return wb
