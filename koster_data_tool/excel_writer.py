from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP

from openpyxl.worksheet.worksheet import Worksheet

from .export_blocks import Block3Header


INT_FMT = "0"
FLOAT2_FMT = "0.00"


def _round_half_up(value: float, ndigits: int) -> float:
    q = Decimal("1") if ndigits == 0 else Decimal("1." + ("0" * ndigits))
    return float(Decimal(str(value)).quantize(q, rounding=ROUND_HALF_UP))


def get_number_format(header: str) -> str | None:
    title = (header or "").strip().lower()
    if "specific capacitance" in title or title == "csp":
        return INT_FMT
    if "specific capacity" in title or title == "qsp":
        return FLOAT2_FMT
    if title in {"r↓", "r_turn", "retention", "ce"}:
        return FLOAT2_FMT
    if "retention" in title:
        return FLOAT2_FMT
    return None


def apply_display_round(value: float, number_format: str | None) -> float:
    if number_format == INT_FMT:
        return _round_half_up(value, 0)
    if number_format == FLOAT2_FMT:
        return _round_half_up(value, 2)
    return value


def write_block(ws: Worksheet, start_col: int, start_row: int, block: Block3Header) -> tuple[int, int]:
    ncol = len(block.h1)
    nrow = max((len(c) for c in block.data), default=0)

    for i in range(ncol):
        c = start_col + i
        ws.cell(row=start_row, column=c, value=block.h1[i] if i < len(block.h1) else "")
        ws.cell(row=start_row + 1, column=c, value=block.h2[i] if i < len(block.h2) else "")
        ws.cell(row=start_row + 2, column=c, value=block.h3[i] if i < len(block.h3) else "")

    for i in range(ncol):
        fmt = get_number_format(block.h1[i] if i < len(block.h1) else "")
        col = block.data[i] if i < len(block.data) else []
        for r, raw_v in enumerate(col):
            cell = ws.cell(row=start_row + 3 + r, column=start_col + i, value=raw_v)
            if isinstance(raw_v, (int, float)) and fmt:
                # 仅用于显示格式，不改变内部计算值。
                _ = apply_display_round(float(raw_v), 0 if fmt == INT_FMT else 2)
                cell.number_format = fmt

    return start_col + ncol, start_row + 3 + nrow


def blank_col_after(ws: Worksheet, col_idx: int) -> int:
    return col_idx + 1
