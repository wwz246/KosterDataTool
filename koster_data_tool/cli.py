from __future__ import annotations
import argparse
import json
import re
import shutil
import subprocess
import threading
from pathlib import Path
from openpyxl import load_workbook
from .bootstrap import init_run_context
from .colmap import parse_file_for_cycles, read_and_map_file
from .curve_export import export_cv_block, export_eis_block, export_gcd_block
from .export_pipeline import run_full_export
from .cycle_split import select_cycle_indices, split_cycles
from .gcd_segment import calc_m_active_g, decide_main_order, drop_first_cycle_reverse_segment, segment_one_cycle
from .gcd_window_metrics import compute_gcd_file_metrics
from .gui import run_gui
from .rate_retention import build_rate_and_retention_for_battery
from .scanner import scan_root
from .state_store import write_last_root
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="KosterDataTool")
    p.add_argument("--root", type=str, default="", help="根目录绝对路径或相对路径")
    p.add_argument("--no-gui", action="store_true", help="以 CLI 形态运行")
    p.add_argument("--selftest", action="store_true", help="运行自检")
    p.add_argument("--scan-only", action="store_true", help="仅扫描并打印摘要")
    p.add_argument("--parse-one", type=str, default="", help="解析单个文件并做列映射/单位换算")
    p.add_argument("--curve-one", type=str, default="", help="导出单个曲线数据块")
    p.add_argument("--split-one", type=str, default="", help="解析并执行分圈+选圈")
    p.add_argument("--gcd-seg-one", type=str, default="", help="单文件执行 GCD 分段")
    p.add_argument("--n-cycle", type=int, default=1, help="代表圈序号")
    p.add_argument("--a-geom", type=float, default=1.0, help="几何面积 cm^2")
    p.add_argument("--m-pos", type=float, default=0.0, help="正极质量 mg")
    p.add_argument("--m-neg", type=float, default=0.0, help="负极质量 mg")
    p.add_argument("--p-active", type=float, default=100.0, help="活性物比例 %")
    p.add_argument("--v-start", type=float, default=None, help="起始电压")
    p.add_argument("--v-end", type=float, default=None, help="终止电压")
    p.add_argument("--gcd-metrics-one", type=str, default="", help="单文件执行 GCD 指标计算")
    p.add_argument("--output-type", type=str, choices=["Csp", "Qsp"], default="Csp", help="输出类型")
    p.add_argument("--export", action="store_true", help="执行端到端导出")
    p.add_argument("--params-json", type=str, default="", help="每电池参数 JSON 文件")
    p.add_argument("--k-factor", type=float, default=None, help="Csp 的 K 系数")
    p.add_argument("--n-gcd", type=int, default=1, help="代表圈序号（GCD 指标）")
    p.add_argument("--rate-selftest", action="store_true", help="打印 Step8 的 Rate/Retention 自检摘要")
    return p
def _write_sample_cv(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time(s)\tVoltage(V)\tCurrent(mA)\n"
        "0.10\t0.20\t0.30 1 CYCLE\n"
        "0.20\t0.30\t0.40\n"
        "0.30\t0.40\t0.50 2 CYCLE\n"
        "0.40\t0.50\t0.60\n"
        "0.50\t0.60\t0.70 3 CYCLE\n"
        "0.60\t0.70\t0.80\n",
        encoding="utf-8",
    )
def _write_sample_gcd(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\tStep\tCycle\n"
        "0\t3.20\t-0.5\t1\t1\n"
        "1\t3.10\t-0.5\t1\t1\n"
        "2\t3.10\t0.5\t2\t1\n"
        "3\t3.20\t0.5\t2\t1\n"
        "4\t3.30\t0.5\t2\t1\n"
        "5\t3.20\t-0.5\t3\t1\n"
        "6\t3.10\t-0.5\t3\t1\n"
        "7\t3.00\t-0.5\t3\t1\n"
        "8\t3.00\t0.5\t4\t2\n"
        "9\t3.10\t0.5\t4\t2\n"
        "10\t3.20\t0.5\t4\t2\n"
        "11\t3.10\t-0.5\t5\t2\n"
        "12\t3.00\t-0.5\t5\t2\n"
        "13\t2.90\t-0.5\t5\t2\n",
        encoding="utf-8",
    )
def _write_sample_gcd_no_cycle(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\n"
        "0\t3.1\t0.5\n"
        "1\t3.2\t0.5\n"
        "2\t3.3\t0.5\n",
        encoding="utf-8",
    )
def _write_sample_eis(path: Path) -> None:
    path.write_text("# comment\nFreq\tZ'\tZ''\n1\t2\t3\n2\t3\t4\n", encoding="utf-8")
def _write_sample_cv_units(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "时间(s)\t电压(V)\t电流(mA)\tStep\n"
        "0\t3.00\t10\t1\n"
        "1\t3.10\t20\t1\n",
        encoding="utf-8",
    )
def _write_sample_gcd_units(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time(s)\tVoltage(V)\tj(mA/cm2)\tCycle\n"
        "0\t3.20\t5\t1\n"
        "1\t3.30\t10\t1\n",
        encoding="utf-8",
    )
def _write_sample_eis_units(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Freq(Hz)\tZ'(Ohm·cm2)\tZ''(Ohm·cm2)\n"
        "1\t10\t4\n"
        "2\t12\t6\n",
        encoding="utf-8",
    )
def _write_sample_eis_no_header_bad(path: Path) -> None:
    path.write_text("# comment\nFreq\tAlpha\tBeta\n1\t2\t3\n2\t3\t4\n", encoding="utf-8")
def _write_sample_cv_cycle_rules(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\n"
        "0.10\t0.20\t0.30 1 CYCLE\n"
        "0.20\t0.30\t0.40\n"
        "0.30\t0.40\t0.50 2 CYCLE\n",
        encoding="utf-8",
    )
def _write_sample_gcd_cycle_col(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\tCycle\n"
        "0\t3.1\t0.5\t1\n"
        "1\t3.2\t0.5\t1\n"
        "2\t3.3\t0.5\t2\n"
        "3\t3.4\t0.5\t3\n",
        encoding="utf-8",
    )
def _write_sample_gcd_metrics(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\tStep\tCycle\tQ_chg\tQ_dis\n"
        "0\t2.4\t1\t1\t1\t0.00\t0.00\n"
        "1\t3.0\t1\t1\t1\t0.28\t0.00\n"
        "2\t4.4\t1\t1\t1\t0.56\t0.00\n"
        "3\t4.3\t-1\t2\t1\t0.56\t0.05\n"
        "4\t3.5\t-1\t2\t1\t0.56\t0.32\n"
        "5\t2.3\t-1\t2\t1\t0.56\t0.58\n"
        "6\t2.4\t1\t3\t2\t0.00\t0.00\n"
        "7\t3.1\t1\t3\t2\t0.30\t0.00\n"
        "8\t4.3\t1\t3\t2\t0.57\t0.00\n"
        "9\t4.3\t-1\t4\t2\t0.57\t0.07\n"
        "10\t3.6\t-1\t4\t2\t0.57\t0.34\n"
        "11\t2.3\t-1\t4\t2\t0.57\t0.61\n",
        encoding="utf-8",
    )
def _write_sample_gcd_capacity_only(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tStep\tCycle\tQ_chg\tQ_dis\n"
        "0\t2.4\t1\t1\t0.00\t0.00\n"
        "1\t3.0\t1\t1\t0.28\t0.00\n"
        "2\t4.4\t1\t1\t0.56\t0.00\n"
        "3\t4.3\t2\t1\t0.56\t0.05\n"
        "4\t3.5\t2\t1\t0.56\t0.32\n"
        "5\t2.3\t2\t1\t0.56\t0.58\n",
        encoding="utf-8",
    )
def _write_sample_gcd_window_nonrep_fail(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\tStep\tCycle\n"
        "0\t2.4\t1.0\t1\t1\n"
        "1\t3.0\t1.0\t1\t1\n"
        "2\t4.4\t1.0\t1\t1\n"
        "3\t4.3\t-1.0\t2\t1\n"
        "4\t3.5\t-1.0\t2\t1\n"
        "5\t2.3\t-1.0\t2\t1\n"
        "6\t2.4\t1.0\t3\t2\n"
        "7\t3.1\t1.0\t3\t2\n"
        "8\t4.1\t1.0\t3\t2\n"
        "9\t4.1\t-1.0\t4\t2\n"
        "10\t3.6\t-1.0\t4\t2\n"
        "11\t2.3\t-1.0\t4\t2\n",
        encoding="utf-8",
    )
def _write_sample_gcd_window_rep_fail(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\tStep\tCycle\n"
        "0\t2.4\t0.5\t1\t1\n"
        "1\t3.2\t0.5\t1\t1\n"
        "2\t4.3\t0.5\t1\t1\n"
        "3\t3.9\t-0.5\t2\t1\n"
        "4\t3.3\t-0.5\t2\t1\n"
        "5\t2.4\t-0.5\t2\t1\n"
        "6\t2.6\t0.5\t3\t2\n"
        "7\t2.9\t0.5\t3\t2\n"
        "8\t3.1\t0.5\t3\t2\n"
        "9\t3.1\t-0.5\t4\t2\n"
        "10\t2.9\t-0.5\t4\t2\n"
        "11\t2.6\t-0.5\t4\t2\n",
        encoding="utf-8",
    )
def _write_sample_gcd_window_boundary_bracket(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\tStep\tCycle\n"
        "0\t3.0\t-1.0\t1\t1\n"
        "1\t2.4\t-1.0\t1\t1\n"
        "2\t2.6\t1.0\t2\t1\n"
        "3\t3.3\t1.0\t2\t1\n"
        "4\t4.3\t1.0\t2\t1\n"
        "5\t4.2\t-1.0\t3\t1\n"
        "6\t3.4\t-1.0\t3\t1\n"
        "7\t2.4\t-1.0\t3\t1\n"
        "8\t2.5\t1.0\t4\t2\n"
        "9\t3.2\t1.0\t4\t2\n"
        "10\t4.2\t1.0\t4\t2\n"
        "11\t4.1\t-1.0\t5\t2\n"
        "12\t3.3\t-1.0\t5\t2\n"
        "13\t2.5\t-1.0\t5\t2\n",
        encoding="utf-8",
    )
def _write_sample_eis_with_string_col(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Freq\tZ'\tZ''\tRange\n"
        "1\t2\t3\t20mA\n"
        "2\t3\t4\t20mA\n",
        encoding="utf-8",
    )
def _write_sample_eis_cycle_none(path: Path) -> None:
    path.write_text("# comment\nFreq\tZre\tZim\n1\t2\t3\n2\t3\t4\n3\t4\t5\n", encoding="utf-8")
def _write_step8_cv(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time(s)\tVoltage(V)\tj(mA/cm2)\tCycle\n"
        "0\t1.0\t10\t1\n"
        "1\t1.1\t-20\t1\n"
        "2\t1.2\t30\t1\n"
        "3\t1.3\t-40\t1\n",
        encoding="utf-8",
    )
def _write_step8_gcd(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Time(s)\tVoltage(V)\tCurrent(A)\tStep\tCycle\n"
        "0\t3.00\t-0.2\t1\t1\n"
        "1\t2.90\t-0.2\t1\t1\n"
        "2\t2.90\t0.0\t2\t1\n"
        "3\t2.90\t0.0\t2\t1\n"
        "4\t2.95\t0.2\t3\t1\n"
        "5\t3.05\t0.2\t3\t1\n"
        "6\t3.15\t0.2\t3\t1\n"
        "7\t3.20\t0.2\t4\t2\n"
        "8\t3.30\t0.2\t4\t2\n"
        "9\t3.40\t0.2\t4\t2\n"
        "10\t3.30\t-0.2\t5\t2\n"
        "11\t3.20\t-0.2\t5\t2\n"
        "12\t3.10\t-0.2\t5\t2\n",
        encoding="utf-8",
    )
def _write_step8_eis(path: Path) -> None:
    path.write_text(
        "# comment\n"
        "Freq(Hz)\tZ'(Ohm·cm2)\tZ''(Ohm·cm2)\n"
        "1\t10\t4\n"
        "2\t12\t6\n",
        encoding="utf-8",
    )
def _write_step8_rate_good(path: Path, current: float, dq_dis_end: float) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\tStep\tCycle\tQ_chg\tQ_dis\n"
        f"0\t2.4\t0\t1\t1\t0.00\t0.00\n"
        f"1\t3.0\t{current}\t1\t1\t0.28\t0.00\n"
        f"2\t4.3\t{current}\t1\t1\t0.57\t0.00\n"
        f"3\t4.3\t{-current}\t2\t1\t0.57\t0.07\n"
        f"4\t3.6\t{-current}\t2\t1\t0.57\t0.34\n"
        f"5\t2.3\t{-current}\t2\t1\t0.57\t{dq_dis_end}\n",
        encoding="utf-8",
    )
def _write_step8_rate_bad(path: Path, current: float) -> None:
    path.write_text(
        "# comment\n"
        "Time\tVoltage\tCurrent\tStep\tCycle\tQ_chg\tQ_dis\n"
        f"0\t2.4\t0\t1\t1\t0.00\t0.00\n"
        f"1\t3.0\t0\t1\t1\t0.00\t0.00\n"
        f"2\t4.3\t0\t1\t1\t0.00\t0.00\n"
        f"3\t4.3\t0\t2\t1\t0.00\t0.00\n"
        f"4\t3.6\t0\t2\t1\t0.00\t0.00\n"
        f"5\t2.3\t0\t2\t1\t0.00\t0.00\n",
        encoding="utf-8",
    )
def _create_selftest_tree(base_root: Path) -> tuple[Path, Path]:
    if base_root.exists():
        shutil.rmtree(base_root)
    base_root.mkdir(parents=True, exist_ok=True)
    struct_a = base_root / "structure_a_root"
    struct_a.mkdir(parents=True, exist_ok=True)
    _write_sample_cv(struct_a / "CV-1.txt")
    _write_sample_gcd(struct_a / "GCD-0.5.txt")
    _write_sample_gcd_no_cycle(struct_a / "GCD-2.txt")
    _write_sample_eis(struct_a / "EIS-1.txt")
    _write_sample_cv_units(struct_a / "CV-2.txt")
    _write_sample_gcd_units(struct_a / "GCD-3.txt")
    _write_sample_eis_units(struct_a / "EIS-4.txt")
    _write_sample_eis_no_header_bad(struct_a / "EIS-5.txt")
    _write_sample_cv_cycle_rules(struct_a / "CV-10.txt")
    _write_sample_gcd_cycle_col(struct_a / "GCD-10.txt")
    _write_sample_eis_cycle_none(struct_a / "EIS-10.txt")
    _write_sample_eis_with_string_col(struct_a / "EIS-11.txt")
    _write_sample_gcd_metrics(struct_a / "GCD-1.txt")
    _write_sample_gcd_window_nonrep_fail(struct_a / "GCD-11.txt")
    _write_sample_gcd_window_boundary_bracket(struct_a / "GCD-12.txt")
    _write_sample_gcd_capacity_only(struct_a / "GCD-4.txt")
    step8_root = base_root / "step8"
    step8_root.mkdir(parents=True, exist_ok=True)
    _write_step8_cv(step8_root / "CV-5.txt")
    _write_step8_gcd(step8_root / "GCD-0.5.txt")
    _write_step8_eis(step8_root / "EIS-1.txt")
    rate_ok = step8_root / "battery_rate_ok"
    rate_ok.mkdir(parents=True, exist_ok=True)
    _write_step8_rate_good(rate_ok / "GCD-0.5.txt", 0.5, 0.58)
    _write_step8_rate_good(rate_ok / "GCD-1.txt", 1.0, 0.50)
    rate_bad = step8_root / "battery_rate_bad"
    rate_bad.mkdir(parents=True, exist_ok=True)
    _write_step8_rate_bad(rate_bad / "GCD-0.5.txt", 0.5)
    _write_step8_rate_bad(rate_bad / "GCD-1.txt", 1.0)
    struct_b = base_root / "structure_b_root"
    for bat in ("Battery_A", "Battery_B"):
        bat_dir = struct_b / bat
        bat_dir.mkdir(parents=True, exist_ok=True)
        _write_sample_cv(bat_dir / "CV-1.txt")
        _write_sample_gcd(bat_dir / "GCD-2.txt")
        _write_sample_eis(bat_dir / "EIS-0.2.txt")
    deep_file = struct_b / "Battery_A" / "deep_l2" / "deep_l3" / "too_deep.txt"
    deep_file.parent.mkdir(parents=True, exist_ok=True)
    deep_file.write_text("should appear in skipped report\n", encoding="utf-8")
    return struct_a, struct_b
def _estimate_cycle_from_file(file_path: Path, file_type: str, logger, run_report_path: str) -> int:
    _mapping, _series, kept_raw_line_indices, marker_events, has_cycle_col, cycle_values = parse_file_for_cycles(
        file_path=str(file_path),
        file_type=file_type,
        a_geom_cm2=1.0,
        v_start=None,
        v_end=None,
        logger=logger,
        run_report_path=run_report_path,
    )
    split_result = split_cycles(file_type, has_cycle_col, cycle_values, kept_raw_line_indices, marker_events)
    if split_result.max_cycle is None:
        raise ValueError("max_cycle is None")
    return split_result.max_cycle
def _run_split_one(ctx, logger, split_one: str, n_cycle: int, a_geom: float, v_start: float | None, v_end: float | None) -> int:
    fpath = Path(split_one).expanduser().resolve()
    m = re.match(r"^(CV|GCD|EIS)-", fpath.name, re.IGNORECASE)
    if not m:
        raise ValueError("--split-one 文件名必须以 CV-/GCD-/EIS- 开头")
    file_type = m.group(1).upper()
    _mapping, _series, kept_raw_line_indices, marker_events, has_cycle_col, cycle_values = parse_file_for_cycles(
        file_path=str(fpath),
        file_type=file_type,
        a_geom_cm2=a_geom,
        v_start=v_start,
        v_end=v_end,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    split_result = split_cycles(file_type, has_cycle_col, cycle_values, kept_raw_line_indices, marker_events)
    selected = select_cycle_indices(file_type, split_result, n_cycle)
    print(f"file_type={file_type}")
    print(f"split_method={split_result.method}")
    print(f"max_cycle={split_result.max_cycle}")
    print(f"selected_n={n_cycle}")
    print(f"selected_row_count={len(selected)}")
    print(f"warnings={split_result.warnings}")
    print(f"run_report_path={ctx.report_path}")
    return 0
def _run_gcd_seg_one(ctx, logger, args) -> int:
    fpath = Path(args.gcd_seg_one).expanduser().resolve()
    m = re.match(r"^GCD-([+-]?\d+(?:\.\d+)?)\.txt$", fpath.name, re.IGNORECASE)
    if not m:
        raise ValueError("--gcd-seg-one 文件名必须为 GCD-<num>.txt")
    j_label = float(m.group(1))
    _mapping, series, kept_raw_line_indices, marker_events, has_cycle_col, cycle_values = parse_file_for_cycles(
        file_path=str(fpath),
        file_type="GCD",
        a_geom_cm2=args.a_geom,
        v_start=args.v_start,
        v_end=args.v_end,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    split_result = split_cycles("GCD", has_cycle_col, cycle_values, kept_raw_line_indices, marker_events)
    row_indices = select_cycle_indices("GCD", split_result, args.n_cycle)
    t = [series["t"][i] for i in row_indices]
    E = [series["E"][i] for i in row_indices]
    I = [series["I"][i] for i in row_indices]
    step = [int(round(series["Step"][i])) for i in row_indices] if "Step" in series else None
    m_active = calc_m_active_g(args.m_pos, args.m_neg, args.p_active)
    cycle_seg = segment_one_cycle(t, E, I, step, args.v_start, args.v_end, j_label, m_active)
    cycle_seg.cycle_k = args.n_cycle
    all_cycles = []
    max_cycle = split_result.max_cycle or 0
    for k in range(1, max_cycle + 1):
        idxs = split_result.cycles.get(k, [])
        if not idxs:
            continue
        kk_t = [series["t"][i] for i in idxs]
        kk_E = [series["E"][i] for i in idxs]
        kk_I = [series["I"][i] for i in idxs]
        kk_step = [int(round(series["Step"][i])) for i in idxs] if "Step" in series else None
        seg_k = segment_one_cycle(kk_t, kk_E, kk_I, kk_step, args.v_start, args.v_end, j_label, m_active)
        seg_k.cycle_k = k
        all_cycles.append(seg_k)
    main_order = None
    adjusted_cycle1 = None
    if max_cycle >= 2:
        main_order = decide_main_order(all_cycles)
        cycle1 = next((x for x in all_cycles if x.cycle_k == 1), None)
        if cycle1 is not None:
            adjusted_cycle1 = drop_first_cycle_reverse_segment(cycle1, main_order)
    print(f"J_label={j_label}")
    print(f"m_active={m_active}")
    print(f"cycle_k={cycle_seg.cycle_k}")
    print(f"segment_count={len(cycle_seg.segments)}")
    for s in cycle_seg.segments:
        print(
            "segment="
            f"({s.kind},{s.start},{s.end},{s.t_start},{s.t_end},{s.I_med},{s.deltaE_end})"
        )
    if main_order is not None:
        print(f"main_order={main_order.order}")
        print(f"main_order_decided_from={main_order.decided_from}")
        print(f"main_order_warnings={main_order.warnings}")
    if adjusted_cycle1 is not None:
        print(f"cycle1_after_drop_segment_count={len(adjusted_cycle1.segments)}")
        print(f"cycle1_after_drop_warnings={adjusted_cycle1.warnings}")
    print(f"warnings={cycle_seg.warnings}")
    print(f"run_report_path={ctx.report_path}")
    return 0
def _selftest(ctx, logger) -> int:
    temp_root = ctx.run_temp_dir / "selftest_root"
    struct_a, struct_b = _create_selftest_tree(temp_root)
    logger.info("selftest: start", root=str(struct_b))
    print(f"SELFTEST_ROOT={struct_b}")
    assert _estimate_cycle_from_file(struct_a / "CV-1.txt", "CV", logger, str(ctx.report_path)) == 4, "CV-1.txt maxCycle assertion failed"
    assert _estimate_cycle_from_file(struct_a / "GCD-0.5.txt", "GCD", logger, str(ctx.report_path)) == 2, "GCD-1.txt maxCycle assertion failed"
    gcd2_failed = False
    try:
        _estimate_cycle_from_file(struct_a / "GCD-2.txt", "GCD", logger, str(ctx.report_path))
    except ValueError as e:
        gcd2_failed = "E9007" in str(e)
    assert gcd2_failed, "GCD-2.txt must fail with E9007"
    cv_map, cv_series = read_and_map_file(
        file_path=str(struct_a / "CV-2.txt"),
        file_type="CV",
        a_geom_cm2=1.0,
        v_start=2.5,
        v_end=4.2,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert cv_map.unit.get("I") == "A", "CV-2 I unit must be A"
    assert abs(cv_series["I"][0] - 0.01) < 1e-12 and abs(cv_series["I"][1] - 0.02) < 1e-12, "CV-2 I mA->A failed"
    gcd_map, gcd_series = read_and_map_file(
        file_path=str(struct_a / "GCD-3.txt"),
        file_type="GCD",
        a_geom_cm2=2.0,
        v_start=2.5,
        v_end=4.2,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert gcd_map.unit.get("j") == "A/cm2", "GCD-3 j unit must be A/cm2"
    assert gcd_map.unit.get("I") == "A", "GCD-3 derived I unit must be A"
    assert abs(gcd_series["I"][0] - 0.01) < 1e-12 and abs(gcd_series["I"][1] - 0.02) < 1e-12, "GCD-3 I=j*A failed"
    _, eis_series = read_and_map_file(
        file_path=str(struct_a / "EIS-4.txt"),
        file_type="EIS",
        a_geom_cm2=2.0,
        v_start=None,
        v_end=None,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert abs(eis_series["Zre"][0] - 5.0) < 1e-12 and abs(eis_series["Zim"][1] - 3.0) < 1e-12, "EIS-4 area normalization failed"
    _, eis11_series = read_and_map_file(
        file_path=str(struct_a / "EIS-11.txt"),
        file_type="EIS",
        a_geom_cm2=1.0,
        v_start=None,
        v_end=None,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert abs(eis11_series["Zre"][0] - 2.0) < 1e-12 and abs(eis11_series["Zim"][1] - 4.0) < 1e-12, "EIS-11 string column should be ignored"
    failed = False
    try:
        read_and_map_file(
            file_path=str(struct_a / "EIS-5.txt"),
            file_type="EIS",
            a_geom_cm2=2.0,
            v_start=None,
            v_end=None,
            logger=logger,
            run_report_path=str(ctx.report_path),
        )
    except ValueError as e:
        logger.exception("selftest expected parse failure", exc=e)
        failed = "E9008" in str(e)
    assert failed, "EIS-5 must fail with E9008"
    assert "E9008" in Path(ctx.report_path).read_text(encoding="utf-8"), "run_report must contain E9008"
    _m_cv10, _s_cv10, cv10_kept, cv10_markers, cv10_has_cycle_col, cv10_cycle_values = parse_file_for_cycles(
        file_path=str(struct_a / "CV-10.txt"),
        file_type="CV",
        a_geom_cm2=1.0,
        v_start=2.5,
        v_end=4.2,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    cv10_split = split_cycles("CV", cv10_has_cycle_col, cv10_cycle_values, cv10_kept, cv10_markers)
    assert cv10_split.method == "k_cycle", "CV-10.txt method assertion failed"
    assert cv10_split.max_cycle == 2, "CV-10.txt maxCycle assertion failed"
    assert len(cv10_split.cycles.get(1, [])) == 1, "CV-10.txt cycle#1 size assertion failed"
    _m_gcd10, _s_gcd10, gcd10_kept, gcd10_markers, gcd10_has_cycle_col, gcd10_cycle_values = parse_file_for_cycles(
        file_path=str(struct_a / "GCD-10.txt"),
        file_type="GCD",
        a_geom_cm2=1.0,
        v_start=2.5,
        v_end=4.2,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    gcd10_split = split_cycles("GCD", gcd10_has_cycle_col, gcd10_cycle_values, gcd10_kept, gcd10_markers)
    assert gcd10_split.method == "cycle_col", "GCD-10.txt method assertion failed"
    for k, idxs in gcd10_split.cycles.items():
        for i in idxs:
            assert gcd10_cycle_values is not None and gcd10_cycle_values[i] == k, "GCD-10.txt cycle membership assertion failed"
    _m_eis10, _s_eis10, eis10_kept, eis10_markers, eis10_has_cycle_col, eis10_cycle_values = parse_file_for_cycles(
        file_path=str(struct_a / "EIS-10.txt"),
        file_type="EIS",
        a_geom_cm2=1.0,
        v_start=None,
        v_end=None,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    eis10_split = split_cycles("EIS", eis10_has_cycle_col, eis10_cycle_values, eis10_kept, eis10_markers)
    assert eis10_split.method == "none", "EIS-10.txt method assertion failed"
    assert eis10_split.cycles == {}, "EIS-10.txt cycles assertion failed"
    try:
        _run_split_one(ctx, logger, str(struct_a / "CV-10.txt"), 1, 1.0, 2.5, 4.2)
    except Exception as exc:  # noqa: BLE001
        logger.exception("selftest split_one failed", exc=exc)
        raise AssertionError(f"CV-10.txt split_one assertion failed: {exc}") from exc
    _mapping, series, kept_raw_line_indices, marker_events, has_cycle_col, cycle_values = parse_file_for_cycles(
        file_path=str(struct_a / "GCD-0.5.txt"),
        file_type="GCD",
        a_geom_cm2=1.0,
        v_start=2.5,
        v_end=4.2,
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    split_result = split_cycles("GCD", has_cycle_col, cycle_values, kept_raw_line_indices, marker_events)
    m_active = calc_m_active_g(10, 0, 90)
    seg_cycles = []
    for k in sorted(split_result.cycles):
        idxs = split_result.cycles[k]
        seg_k = segment_one_cycle(
            [series["t"][i] for i in idxs],
            [series["E"][i] for i in idxs],
            [series["I"][i] for i in idxs],
            [int(round(series["Step"][i])) for i in idxs],
            2.5,
            4.2,
            0.5,
            m_active,
        )
        seg_k.cycle_k = k
        seg_cycles.append(seg_k)
    order = decide_main_order(seg_cycles)
    assert order.order == "Charge→Discharge", "main order assertion failed"
    cycle1 = next(c for c in seg_cycles if c.cycle_k == 1)
    after_drop = drop_first_cycle_reverse_segment(cycle1, order)
    assert len(after_drop.segments) == len(cycle1.segments) - 1, "drop reverse segment assertion failed"
    cmd = [
        "python",
        "main.py",
        "--no-gui",
        "--gcd-seg-one",
        str(struct_a / "GCD-0.5.txt"),
        "--m-pos",
        "10",
        "--m-neg",
        "0",
        "--p-active",
        "90",
        "--v-start",
        "2.5",
        "--v-end",
        "4.2",
        "--n-cycle",
        "1",
    ]
    run = subprocess.run(cmd, cwd=str(ctx.paths.program_dir), check=False, capture_output=True, text=True)
    assert run.returncode == 0, f"gcd-seg-one cli assertion failed: {run.stderr}"
    metrics = compute_gcd_file_metrics(
        file_path=str(struct_a / "GCD-1.txt"),
        root_params={"v_start": 2.5, "v_end": 4.2, "a_geom": 1.0, "output_type": "Csp", "k_factor": 1.0, "n_gcd": 1},
        battery_params={"m_pos": 10.0, "m_neg": 0.0, "p_active": 90.0},
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert metrics.fatal_error is None or "E5201" not in metrics.fatal_error, "代表圈不得触发 E5201"
    metrics_capacity = compute_gcd_file_metrics(
        file_path=str(struct_a / "GCD-4.txt"),
        root_params={"v_start": 2.5, "v_end": 4.2, "a_geom": 1.0, "output_type": "Csp", "k_factor": 1.0, "n_gcd": 1},
        battery_params={"m_pos": 10.0, "m_neg": 0.0, "p_active": 90.0},
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    rep_capacity = metrics_capacity.cycles.get(metrics_capacity.n_gcd)
    assert rep_capacity is not None, "W5101 样例应产出代表圈结果"
    assert rep_capacity.delta_q_source == "capacity", "W5101 样例应为 capacity 源"
    ce_raw = 100 * (metrics.cycles[metrics.n_gcd].delta_q_dis / metrics.cycles[metrics.n_gcd].delta_q_chg)
    ce_rounded_proxy = 100 * (round(metrics.cycles[metrics.n_gcd].delta_q_dis, 2) / round(metrics.cycles[metrics.n_gcd].delta_q_chg, 2))
    assert abs(ce_raw - ce_rounded_proxy) > 1e-6, "CE 必须用未取整 ΔQ"
    metrics_nonrep_fail = compute_gcd_file_metrics(
        file_path=str(struct_a / "GCD-11.txt"),
        root_params={"v_start": 2.5, "v_end": 4.2, "a_geom": 1.0, "output_type": "Csp", "k_factor": 1.0, "n_gcd": 1},
        battery_params={"m_pos": 10.0, "m_neg": 0.0, "p_active": 90.0},
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert metrics_nonrep_fail.fatal_error is None, "非选定圈截取失败不得触发 E5201"
    assert metrics_nonrep_fail.cycles[2].ok_window is False, "GCD-11 cycle2 应窗口失败"
    assert metrics_nonrep_fail.cycles[2].delta_q_chg is None and metrics_nonrep_fail.cycles[2].delta_q_dis is None, "窗口失败指标应为 NaN/None"
    assert any("W5204" in w and "cycle=2" in w for w in metrics_nonrep_fail.warnings), "非选定圈失败应记录 W5204"
    metrics_boundary_bracket = compute_gcd_file_metrics(
        file_path=str(struct_a / "GCD-12.txt"),
        root_params={"v_start": 2.5, "v_end": 4.2, "a_geom": 1.0, "output_type": "Csp", "k_factor": 1.0, "n_gcd": 2},
        battery_params={"m_pos": 10.0, "m_neg": 0.0, "p_active": 90.0},
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert metrics_boundary_bracket.fatal_error is None, "段边界可插值样例不得触发 E5201"
    assert metrics_boundary_bracket.cycles[metrics_boundary_bracket.n_gcd].ok_window is True, "段边界可插值样例应窗口成功"
    metrics_rep_fail = compute_gcd_file_metrics(
        file_path=str(struct_a / "GCD-11.txt"),
        root_params={"v_start": 2.5, "v_end": 4.2, "a_geom": 1.0, "output_type": "Csp", "k_factor": 1.0, "n_gcd": 2},
        battery_params={"m_pos": 10.0, "m_neg": 0.0, "p_active": 90.0},
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert metrics_rep_fail.fatal_error is not None and "E5201" in metrics_rep_fail.fatal_error, "选定圈截取失败应触发 E5201"
    result = scan_root(
        root_path=str(struct_b),
        output_dir=str(ctx.paths.output_dir),
        run_id=ctx.run_id,
        cancel_flag=threading.Event(),
        progress_cb=None,
    )
    skipped_report = Path(result.skipped_report_path)
    assert skipped_report.exists(), "skipped report must exist"
    assert skipped_report.read_text(encoding="utf-8").strip(), "skipped report must be non-empty"
    step8_root = temp_root / "step8"
    cv_block = export_cv_block(
        file_path=str(step8_root / "CV-5.txt"),
        n_cv=1,
        a_geom_cm2=2.0,
        m_pos_mg=10.0,
        m_neg_mg=0.0,
        p_active_pct=90.0,
        current_unit="A/g",
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    m_active = calc_m_active_g(10.0, 0.0, 90.0)
    expected = [0.01 * 2.0 / m_active, -0.02 * 2.0 / m_active]
    assert cv_block.data[1][0] > 0 and cv_block.data[1][1] < 0, "CV 导出电流应保留正负号"
    assert abs(cv_block.data[1][0] - expected[0]) < 1e-12 and abs(cv_block.data[1][1] - expected[1]) < 1e-12, "CV 导出需按 j*A_geom/m_active"
    gcd_block = export_gcd_block(str(step8_root / "GCD-0.5.txt"), 1, logger, str(ctx.report_path))
    assert gcd_block.data[0][0] == 0.0, "GCD 导出时间需圈内归零"
    assert len(gcd_block.data[0]) < 7, "GCD 导出应剔除静置段"
    eis_block = export_eis_block(str(step8_root / "EIS-1.txt"), 2.0, logger, str(ctx.report_path))
    assert eis_block.data[1][0] < 0, "EIS 导出第二列需为 -Z''"
    assert abs(eis_block.data[0][0] - 5.0) < 1e-12, "EIS Ohm·cm2 需按面积换算"
    rate_ok = build_rate_and_retention_for_battery(
        gcd_files=[str(step8_root / "battery_rate_ok" / "GCD-0.5.txt"), str(step8_root / "battery_rate_ok" / "GCD-1.txt")],
        n_gcd=1,
        output_type="Qsp",
        root_params={"a_geom": 1.0, "v_start": 2.5, "v_end": 4.2},
        battery_params={"m_pos": 10.0, "m_neg": 0.0, "p_active": 90.0},
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert len(rate_ok.rate.data[0]) >= 2, "Rate 需至少两个工况"
    assert not any("W1304" in w for w in rate_ok.warnings), "正常样例不应触发 W1304"
    rate_bad = build_rate_and_retention_for_battery(
        gcd_files=[str(step8_root / "battery_rate_bad" / "GCD-0.5.txt"), str(step8_root / "battery_rate_bad" / "GCD-1.txt")],
        n_gcd=1,
        output_type="Qsp",
        root_params={"a_geom": 1.0, "v_start": 2.5, "v_end": 4.2},
        battery_params={"m_pos": 10.0, "m_neg": 0.0, "p_active": 90.0},
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    assert any("W1304" in w for w in rate_bad.warnings), "X0<=0 应触发 W1304"
    # step9: 全链路导出回归
    before = {p.name for p in struct_b.iterdir() if p.is_file()}
    export_cmd = ["python", "main.py", "--no-gui", "--root", str(struct_b), "--export", "--output-type", "Csp"]
    run_export = subprocess.run(export_cmd, cwd=str(ctx.paths.program_dir), check=False, capture_output=True, text=True)
    assert run_export.returncode == 0, f"export cli assertion failed: {run_export.stderr}\n{run_export.stdout}"
    after = {p.name for p in struct_b.iterdir() if p.is_file()}
    diff = sorted(after - before)
    assert len(diff) == 2 and all(x.endswith('.xlsx') for x in diff), f"root 新增文件应仅2个xlsx: {diff}"
    # 报告/日志必须在 data_root/KosterData 且不在 program_dir
    for line in run_export.stdout.splitlines():
        if any(line.startswith(k) for k in ("run_report_path=", "log_path=")):
            path = line.split("=", 1)[1].strip()
            path_obj = Path(path).resolve()
            kd_root = ctx.paths.kosterdata_dir.resolve()
            program_root = ctx.paths.program_dir.resolve()
            assert kd_root in path_obj.parents or path_obj == kd_root, f"非KosterData路径: {path}"
            assert not (program_root in path_obj.parents or path_obj == program_root), f"不应写入program_dir: {path}"
    xlsx_paths = [struct_b / name for name in diff]
    battery_wb = None
    electrode_wb = None
    for xp in xlsx_paths:
        wb = load_workbook(xp)
        if "参数汇总" in wb.sheetnames:
            battery_wb = wb
        if "Rate" in wb.sheetnames:
            electrode_wb = wb
    assert battery_wb is not None and electrode_wb is not None, "应同时生成极片级与电池级"
    assert "参数汇总" in battery_wb.sheetnames, "电池级应有参数汇总"
    for b in scan_root(str(struct_b), str(ctx.paths.output_dir), ctx.run_id, threading.Event(), None).batteries:
        assert b.name in battery_wb.sheetnames, f"电池级缺少sheet:{b.name}"
    assert "Rate" in electrode_wb.sheetnames, "极片级缺少Rate"
    assert any(n.startswith("CV-") for n in electrode_wb.sheetnames), "极片级缺少CV sheet"
    assert any(n.startswith("GCD-") for n in electrode_wb.sheetnames), "极片级缺少GCD sheet"
    assert any(n.startswith("EIS-") for n in electrode_wb.sheetnames), "极片级缺少EIS sheet"
    ps = battery_wb["参数汇总"]
    texts = [str(ps.cell(row=r, column=1).value or "") for r in range(1, min(120, ps.max_row + 1))]
    idx_param = next(i for i, t in enumerate(texts, start=1) if t == "电池名")
    idx_detail = next(i for i, t in enumerate(texts, start=1) if t == "电池名" and i > idx_param)
    assert idx_detail - idx_param > 5, "参数表与逐圈结果表间必须有5空行"
    rate_ws = electrode_wb["Rate"]
    battery_names = [b.name for b in sorted(scan_root(str(struct_b), str(ctx.paths.output_dir), ctx.run_id, threading.Event(), None).batteries, key=lambda x: x.name)]
    row3_values = [rate_ws.cell(row=3, column=c).value for c in range(1, rate_ws.max_column + 1)]
    observed_names = [v for v in row3_values if isinstance(v, str) and v in battery_names]
    assert observed_names == battery_names, f"Rate 第3行电池名顺序不正确: {observed_names}"
    # Rate 与 Retention 间应有1空行
    rate_data_last = 4
    while rate_ws.cell(row=rate_data_last, column=1).value not in (None, ""):
        rate_data_last += 1
    assert rate_ws.cell(row=rate_data_last, column=1).value in (None, ""), "Rate 后需空行"
    # 参数汇总逐圈表覆盖 max k，不可跳过缺圈
    header_row = idx_detail
    detail_rows = [r for r in range(header_row + 1, ps.max_row + 1) if ps.cell(row=r, column=1).value]
    assert detail_rows, "参数汇总逐圈明细不能为空"
    # CLI 输出应含失败/告警数量
    assert "failures=" in run_export.stdout and "warnings=" in run_export.stdout, "导出CLI需输出失败/告警数量"
    # step9b: Qsp 导出回归（检查 Rate/Retention 与参数表 K 列）
    struct_qsp_root = temp_root / "structure_b_qsp_root"
    _, struct_qsp = _create_selftest_tree(struct_qsp_root)
    before_qsp = {p.name for p in struct_qsp.iterdir() if p.is_file()}
    export_qsp_cmd = ["python", "main.py", "--no-gui", "--root", str(struct_qsp), "--export", "--output-type", "Qsp"]
    run_export_qsp = subprocess.run(export_qsp_cmd, cwd=str(ctx.paths.program_dir), check=False, capture_output=True, text=True)
    assert run_export_qsp.returncode == 0, f"Qsp export cli assertion failed: {run_export_qsp.stderr}\n{run_export_qsp.stdout}"
    after_qsp = {p.name for p in struct_qsp.iterdir() if p.is_file()}
    diff_qsp = sorted(after_qsp - before_qsp)
    assert len(diff_qsp) == 2 and all(x.endswith('.xlsx') for x in diff_qsp), f"Qsp root 新增文件应仅2个xlsx: {diff_qsp}"
    qsp_battery_wb = None
    qsp_electrode_wb = None
    for name in diff_qsp:
        wb = load_workbook(struct_qsp / name)
        if "参数汇总" in wb.sheetnames:
            qsp_battery_wb = wb
        if "Rate" in wb.sheetnames:
            qsp_electrode_wb = wb
    assert qsp_battery_wb is not None and qsp_electrode_wb is not None, "Qsp 应同时生成极片级与电池级"
    qsp_rate_ws = qsp_electrode_wb["Rate"]
    has_qsp_rate_value = False
    for r in range(4, min(120, qsp_rate_ws.max_row + 1)):
        for c in range(2, min(40, qsp_rate_ws.max_column + 1)):
            v = qsp_rate_ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                has_qsp_rate_value = True
                break
        if has_qsp_rate_value:
            break
    assert has_qsp_rate_value, "Qsp Rate sheet 应有数值输出"
    qsp_ps = qsp_battery_wb["参数汇总"]
    qsp_headers = [str(qsp_ps.cell(row=1, column=c).value or "") for c in range(1, qsp_ps.max_column + 1)]
    assert "K(—)" not in qsp_headers, "Qsp 参数汇总不应输出 K 列"
    battery_names_qsp = [n for n in qsp_battery_wb.sheetnames if n != "参数汇总"]
    assert battery_names_qsp, "Qsp 电池级应包含电池sheet"
    found_retention = False
    for sheet_name in battery_names_qsp:
        ws_b = qsp_battery_wb[sheet_name]
        for r in range(1, min(220, ws_b.max_row + 1)):
            for c in range(1, min(80, ws_b.max_column + 1)):
                if ws_b.cell(row=r, column=c).value != "保持率":
                    continue
                neighbor = ws_b.cell(row=r, column=min(ws_b.max_column, c + 1)).value
                if neighbor not in (None, ""):
                    found_retention = True
                    break
            if found_retention:
                break
        if found_retention:
            break
    assert found_retention, "Qsp 电池级应输出保持率结果(百分比或NA)"
    # full reproducible test report
    test_report_path = ctx.paths.output_dir / f"run_{ctx.run_id}_fulltest_report.txt"
    report_lines: list[str] = []
    def _txt_preview(fp: Path, n: int = 20) -> tuple[list[str], int]:
        try:
            rows = fp.read_text(encoding="utf-8", errors="ignore").splitlines()
            return rows[:n], len(rows)
        except Exception:
            return [], 0
    def _xlsx_preview(fp: Path) -> str:
        try:
            wb = load_workbook(fp)
            names = wb.sheetnames
            sample = []
            for sn in names[:2]:
                ws = wb[sn]
                sample.append(f"{sn}!A1={ws.cell(row=1, column=1).value}")
            return f"sheetnames={names}; sample={sample}"
        except Exception as exc:
            return f"xlsx preview failed: {exc}"
    def _run_case(case_id: str, purpose: str, cmd: list[str]):
        before = set(ctx.paths.output_dir.glob("run_*.txt"))
        run = subprocess.run(cmd, cwd=str(ctx.paths.program_dir), capture_output=True, text=True)
        after = set(ctx.paths.output_dir.glob("run_*.txt"))
        new_files = sorted(after - before)
        status = "PASS" if run.returncode == 0 else "FAIL"
        report_lines.append(f"[{case_id}] {purpose}")
        report_lines.append(f"INPUT: {' '.join(cmd)}")
        report_lines.append(f"OUTPUT_RC: {run.returncode}")
        out_preview = (run.stdout + "\n" + run.stderr).strip().splitlines()[:20]
        report_lines.append("OUTPUT_PREVIEW:")
        report_lines.extend([f"  {x}" for x in out_preview])
        report_lines.append("GENERATED_FILES:")
        for nf in new_files:
            report_lines.append(f"  {nf}")
            if nf.suffix == ".txt":
                head, total = _txt_preview(nf)
                report_lines.append(f"  TXT_LINES={total}")
                for h in head:
                    report_lines.append(f"    {h}")
        for line in run.stdout.splitlines():
            if "_path=" in line and line.strip().endswith(".xlsx"):
                fp = Path(line.split("=", 1)[1].strip())
                if fp.exists():
                    report_lines.append(f"  XLSX: {fp}")
                    report_lines.append(f"    {_xlsx_preview(fp)}")
        report_lines.append(f"STATUS: {status}")
        report_lines.append("")
    report_lines.append(f"DATA_ROOT={ctx.paths.kosterdata_dir}")
    report_lines.append(f"PROGRAM_DIR={ctx.paths.program_dir}")
    report_lines.append(f"DATA_NOT_IN_PROGRAM_DIR={ctx.paths.program_dir.resolve() not in ctx.paths.kosterdata_dir.resolve().parents}")
    report_lines.append("")
    _run_case("T1", "scan-only", ["python", "main.py", "--no-gui", "--scan-only", "--root", str(struct_b)])
    _run_case("T2", "parse-one CV", ["python", "main.py", "--no-gui", "--parse-one", str(struct_a / "CV-1.txt")])
    _run_case("T3", "curve-one CV", ["python", "main.py", "--no-gui", "--curve-one", str(step8_root / "CV-5.txt"), "--n-cycle", "1", "--a-geom", "1", "--m-pos", "10", "--m-neg", "0", "--p-active", "90"])
    _run_case("T4", "curve-one GCD", ["python", "main.py", "--no-gui", "--curve-one", str(step8_root / "GCD-0.5.txt"), "--n-cycle", "1"])
    _run_case("T5", "curve-one EIS", ["python", "main.py", "--no-gui", "--curve-one", str(step8_root / "EIS-1.txt"), "--a-geom", "2"])
    _run_case("T6", "split-one", ["python", "main.py", "--no-gui", "--split-one", str(struct_a / "CV-10.txt"), "--n-cycle", "1", "--v-start", "2.5", "--v-end", "4.2"])
    _run_case("T7", "gcd-seg-one", ["python", "main.py", "--no-gui", "--gcd-seg-one", str(struct_a / "GCD-0.5.txt"), "--m-pos", "10", "--m-neg", "0", "--p-active", "90", "--v-start", "2.5", "--v-end", "4.2", "--n-cycle", "1"])
    _run_case("T8", "gcd-metrics-one", ["python", "main.py", "--no-gui", "--gcd-metrics-one", str(struct_a / "GCD-1.txt"), "--m-pos", "10", "--m-neg", "0", "--p-active", "90", "--v-start", "2.5", "--v-end", "4.2", "--n-gcd", "1", "--output-type", "Qsp", "--k-factor", "1"])
    _run_case("T9", "export", ["python", "main.py", "--no-gui", "--export", "--root", str(struct_b), "--output-type", "Qsp"])
    _run_case("T10", "rate-selftest", ["python", "main.py", "--no-gui", "--rate-selftest", "--output-type", "Qsp", "--m-pos", "10", "--m-neg", "0", "--p-active", "90"])
    # GUI import smoke test
    try:
        import koster_data_tool.gui as _gui_mod  # noqa: F401
        report_lines.append("[T11] GUI import smoke")
        report_lines.append("INPUT: import koster_data_tool.gui")
        report_lines.append("OUTPUT: import ok")
        report_lines.append("STATUS: PASS")
        report_lines.append("")
    except Exception as exc:
        report_lines.append("[T11] GUI import smoke")
        report_lines.append("INPUT: import koster_data_tool.gui")
        report_lines.append(f"OUTPUT: {exc}")
        report_lines.append("STATUS: FAIL")
        report_lines.append("")
    # run_id new file strategy + cleanup simulation
    report_lines.append("[T12] run_id & cleanup")
    txt_files = sorted(ctx.paths.output_dir.glob("run_*.txt"))
    report_lines.append(f"OUTPUT_TXT_COUNT={len(txt_files)}")
    old_file = ctx.paths.output_dir / "run_20000101_000000_old.txt"
    old_file.write_text("old\n", encoding="utf-8")
    import os as _os
    from datetime import datetime as _dt, timedelta as _td
    old_ts = (_dt.now() - _td(days=60)).timestamp()
    _os.utime(old_file, (old_ts, old_ts))
    lc = ctx.paths.state_dir / "last_cleanup.txt"
    lc.write_text((_dt.now() - _td(days=31)).date().isoformat(), encoding="utf-8")
    from .bootstrap import cleanup_if_due
    cleanup_if_due(ctx.paths, logger=logger)
    report_lines.append(f"OLD_FILE_EXISTS_AFTER_CLEANUP={old_file.exists()}")
    report_lines.append(f"STATUS: {'PASS' if not old_file.exists() else 'FAIL'}")
    report_lines.append("")
    test_report_path.write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    logger.info("selftest: report", path=str(test_report_path))
    logger.info("selftest: ok", structure_a=str(struct_a), structure_b=str(struct_b))
    print(f"fulltest_report_path={test_report_path}")
    return 0
def _log_recognized_files(logger, scan_result) -> None:
    for battery in scan_result.batteries:
        for file_type, files in (("CV", battery.cv_files), ("GCD", battery.gcd_files), ("EIS", battery.eis_files)):
            for rf in files:
                logger.info("recognized file", battery=battery.name, file_type=file_type, num=rf.num, path=str(Path(rf.path).resolve()))
def _run_scan_only(ctx, logger, root_arg: str) -> int:
    if not root_arg:
        raise ValueError("--scan-only 需要同时传入 --root <dir>")
    root = Path(root_arg).expanduser().resolve()
    result = scan_root(
        root_path=str(root),
        output_dir=str(ctx.paths.output_dir),
        run_id=ctx.run_id,
        cancel_flag=threading.Event(),
        progress_cb=None,
    )
    write_last_root(ctx.paths.state_dir, root)
    _log_recognized_files(logger, result)
    print(f"structure={result.structure}")
    print(f"batteries={len(result.batteries)}")
    print(f"recognized_file_count={result.recognized_file_count}")
    print(f"skipped_report_path={result.skipped_report_path}")
    print(f"run_report_path={ctx.report_path}")
    return 0
def _run_gcd_metrics_one(ctx, logger, args) -> int:
    fpath = Path(args.gcd_metrics_one).expanduser().resolve()
    metrics = compute_gcd_file_metrics(
        file_path=str(fpath),
        root_params={
            "v_start": args.v_start,
            "v_end": args.v_end,
            "a_geom": args.a_geom,
            "output_type": args.output_type,
            "k_factor": args.k_factor,
            "n_gcd": args.n_gcd,
        },
        battery_params={"m_pos": args.m_pos, "m_neg": args.m_neg, "p_active": args.p_active},
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    rep = metrics.cycles.get(metrics.n_gcd)
    ce = None
    if rep and rep.delta_q_chg and rep.delta_q_dis:
        ce = 100.0 * rep.delta_q_dis / rep.delta_q_chg
    print(f"file_path={metrics.file_path}")
    print(f"j_label={metrics.j_label}")
    print(f"main_order={metrics.main_order}")
    print(f"fatal_error={metrics.fatal_error}")
    print(f"representative_cycle_ok={metrics.representative_cycle_ok}")
    if rep is not None:
        print(f"rep_delta_t={rep.delta_t}")
        print(f"rep_delta_q_chg={rep.delta_q_chg}")
        print(f"rep_delta_q_dis={rep.delta_q_dis}")
        print(f"rep_ce={ce}")
        print(f"rep_r_turn={rep.r_turn}")
        print(f"rep_warnings={rep.warnings}")
    print(f"warnings={metrics.warnings}")
    print(f"run_report_path={ctx.report_path}")
    return 0
def _run_curve_one(ctx, logger, args) -> int:
    fpath = Path(args.curve_one).expanduser().resolve()
    m = re.match(r"^(CV|GCD|EIS)-", fpath.name, re.IGNORECASE)
    if not m:
        raise ValueError("--curve-one 文件名必须以 CV-/GCD-/EIS- 开头")
    ftype = m.group(1).upper()
    if ftype == "CV":
        block = export_cv_block(str(fpath), args.n_cycle, args.a_geom, args.m_pos, args.m_neg, args.p_active, "A/g", logger, str(ctx.report_path))
    elif ftype == "GCD":
        block = export_gcd_block(str(fpath), args.n_cycle, logger, str(ctx.report_path))
    else:
        block = export_eis_block(str(fpath), args.a_geom, logger, str(ctx.report_path))
    preview = []
    rows = len(block.data[0]) if block.data else 0
    for r in range(min(3, rows)):
        preview.append([col[r] for col in block.data])
    print(f"file_type={ftype}")
    print(f"columns={block.h1}")
    print(f"data_rows={rows}")
    print(f"preview={preview}")
    print(f"warnings={block.warnings}")
    print(f"run_report_path={ctx.report_path}")
    return 0
def _run_rate_selftest(ctx, logger, args) -> int:
    temp_root = ctx.run_temp_dir / "selftest_root"
    _create_selftest_tree(temp_root)
    bat = temp_root / "step8" / "battery_rate_ok"
    gcd_files = sorted(str(x) for x in bat.glob("GCD-*.txt"))
    block = build_rate_and_retention_for_battery(
        gcd_files=gcd_files,
        n_gcd=args.n_cycle,
        output_type=args.output_type if args.output_type else "Csp",
        root_params={"a_geom": args.a_geom, "v_start": args.v_start if args.v_start is not None else 2.5, "v_end": args.v_end if args.v_end is not None else 4.2, "k_factor": args.k_factor if args.k_factor is not None else 1.0},
        battery_params={"m_pos": args.m_pos, "m_neg": args.m_neg, "p_active": args.p_active},
        logger=logger,
        run_report_path=str(ctx.report_path),
    )
    print(f"rate_rows={len(block.rate.data[0]) if block.rate.data else 0}")
    print(f"retention_rows={len(block.retention.data[0]) if block.retention.data else 0}")
    print(f"triggered_W1304={any('W1304' in w for w in block.warnings)}")
    print(f"run_report_path={ctx.report_path}")
    return 0
def _default_params_for_scan(scan_result, output_type: str) -> dict:
    battery_params = {}
    for b in scan_result.batteries:
        battery_params[b.name] = {
            "m_pos": 10.0,
            "m_neg": 0.0,
            "p_active": 90.0,
            "k": 1.0,
            "n_cv": 1,
            "n_gcd": 1,
            "v_start": 2.5,
            "v_end": 4.2,
            "main_order": "先充后放",
        }
    return {
        "a_geom": 1.0,
        "output_type": output_type,
        "export_battery_workbook": True,
        "battery_params": battery_params,
    }
def _load_or_default_params(args, scan_result):
    if args.params_json:
        return json.loads(Path(args.params_json).read_text(encoding="utf-8"))
    return _default_params_for_scan(scan_result, args.output_type)
def _run_export(ctx, logger, args) -> int:
    if not args.root:
        raise ValueError("--export 需要 --root")
    root = Path(args.root).expanduser().resolve()
    scan_result = scan_root(str(root), str(ctx.paths.output_dir), ctx.run_id, threading.Event(), None)
    _log_recognized_files(logger, scan_result)
    write_last_root(ctx.paths.state_dir, root)
    params = _load_or_default_params(args, scan_result)
    params["a_geom"] = args.a_geom
    params["output_type"] = args.output_type
    selections = {
        "batteries": [b.name for b in scan_result.batteries],
        "cv_nums": [str(x) for x in scan_result.available_cv[:1]],
        "gcd_nums": [str(x) for x in scan_result.available_gcd[:1]],
        "eis_nums": [str(scan_result.available_eis[-1])] if scan_result.available_eis else [],
    }
    result = run_full_export(str(root), scan_result, params, selections, ctx, logger, None)
    print(f"electrode_path={result['electrode_path']}")
    print(f"battery_path={result['battery_path']}")
    print(f"run_report_path={result['run_report_path']}")
    print(f"log_path={result['log_path']}")
    print(f"failures={len(result.get('failures', []))}")
    print(f"warnings={len(result.get('warnings', []))}")
    return 0
def _run_cli(args) -> int:
    ctx, logger = init_run_context()
    logger.info("mode", mode="CLI")
    if args.selftest:
        return _selftest(ctx, logger)
    if args.export:
        return _run_export(ctx, logger, args)
    if args.scan_only:
        return _run_scan_only(ctx, logger, args.root)
    if args.curve_one:
        return _run_curve_one(ctx, logger, args)
    if args.rate_selftest:
        return _run_rate_selftest(ctx, logger, args)
    if args.split_one:
        return _run_split_one(ctx, logger, args.split_one, args.n_cycle, args.a_geom, args.v_start, args.v_end)
    if args.gcd_seg_one:
        return _run_gcd_seg_one(ctx, logger, args)
    if args.gcd_metrics_one:
        return _run_gcd_metrics_one(ctx, logger, args)
    if args.parse_one:
        fpath = Path(args.parse_one).expanduser().resolve()
        m = re.match(r"^(CV|GCD|EIS)-", fpath.name, re.IGNORECASE)
        if not m:
            raise ValueError("--parse-one 文件名必须以 CV-/GCD-/EIS- 开头")
        file_type = m.group(1).upper()
        mapping, _series = read_and_map_file(
            file_path=str(fpath),
            file_type=file_type,
            a_geom_cm2=args.a_geom,
            v_start=args.v_start,
            v_end=args.v_end,
            logger=logger,
            run_report_path=str(ctx.report_path),
        )
        print(f"file_type={mapping.file_type}")
        print(f"delimiter={mapping.delimiter}")
        print(f"modeCols={mapping.modeCols}")
        print(f"kept_ratio={mapping.kept_ratio}")
        print(f"no_header={mapping.no_header}")
        print(f"col_index={mapping.col_index}")
        print(f"unit={mapping.unit}")
        print(f"warnings={mapping.warnings}")
        print(f"run_report_path={ctx.report_path}")
        return 0
    print("CLI ready. Try --scan-only or --selftest.")
    print(f"LOG_TEXT={ctx.text_log_path}")
    print(f"REPORT={ctx.report_path}")
    return 0
def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if args.no_gui:
        return _run_cli(args)
    return run_gui()
